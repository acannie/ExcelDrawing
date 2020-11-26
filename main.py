import numpy as np
import openpyxl
import math
import sys
from PIL import Image
from openpyxl.styles import PatternFill

# 定数
N = 3  # 1つのセルの縦幅に対する横幅の大きさ
R = 0
G = 1
B = 2

# ファイルのインポート
input_file = 'icon.jpg'
im = np.array(Image.open('figure/' + input_file))

# 画像のサイズを取得
IMG_HEIGHT = im.shape[0]
IMG_WIDTH = im.shape[1]

COLOR_VARIETY = im.shape[2]

# ワークブックを作成
wb = openpyxl.Workbook()
ws = wb.worksheets[0]

# RGBの数値が1桁の場合冒頭に0をつけて2桁にして返す
def addZero(strHexNum='0xAA'):
    if len(strHexNum) == 3:
        return '0x0' + strHexNum[2]
    elif len(strHexNum) == 4:
        return strHexNum
    else:
        print('error!')
        sys.exit()

# RGBをカラーコードに変換
def RGBtoColorCode(RGB=[0, 0, 0]):
    color_code = ''
    for color in range(COLOR_VARIETY):
        color_code += str(addZero(hex(RGB[color])))
    color_code = color_code.replace('0x', '') # 0xrr0xgg0xbb を rrggbb の形にする
    return color_code

# ゲージ
GAUGE_WIDTH = 30
now_pacentage = 0 # 範囲: 0 - 100

print('0', end='')
for g in range(GAUGE_WIDTH):
    print(' ', end='')
print('100%')

print(' ', end='')

# 出力
for i in range(0, IMG_HEIGHT):
    for j in range(0, math.floor(IMG_WIDTH / N)):
        # 横に並んだ N ピクセルの平均 RGB をセルの背景色とする
        ave_RGB = [0, 0, 0]

        for n in range(N):
            for color in range(COLOR_VARIETY):
                ave_RGB[color] += im[i][N * j + n][color]

        for color in range(COLOR_VARIETY):
            ave_RGB[color] = round(ave_RGB[color] / N)

        fill = PatternFill(patternType='solid', fgColor=RGBtoColorCode(ave_RGB))
        ws.cell(row=i+1, column=j+1).fill = fill

    # ゲージを増やす
    if math.ceil(i / IMG_HEIGHT * GAUGE_WIDTH) > now_pacentage:
        print("■", end="", flush=True)
        now_pacentage = math.ceil(i / IMG_HEIGHT * GAUGE_WIDTH)

print('')

# ファイルのエクスポート
print("saving...")

output_file = input_file.replace('.jpg', '.xlsx') # 元画像名と同名の .xlsx ファイルとして出力
wb.save('result/' + output_file)
wb.close()

print("complete!")
